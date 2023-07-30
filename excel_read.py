#!/usr/bin/python3
# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class CustomError(Exception):
    pass


# Find the column based on the value in the first row
def find_column_by_value(sheet, value):
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == value:
                return cell.column_letter

    return None


# Find rows based on column and value
def find_rows_by_value(sheet, column, value, accept_empty=False):
    column_index = column_index_from_string(column) - 1  # Convert to column index
    rows = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cell_value = row[column_index]
        if cell_value == value:
            rows.append(row_index)
        elif cell_value is None:
            if accept_empty:
                rows.append(row_index)

    return rows


def process_excel_file(file_path=None, sheet=None, return_info=False, print_info=False):
    if file_path is None:
        file_path = 'SSH.xlsx'
    if sheet is None:
        sheet = 'Sheet1'
    info_to_return = []

    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet]
    except Exception as e:
        raise CustomError('Error reading the file')

    headers_to_find = ['Hostname', 'Address', 'Username']
    header_dict = {}

    # Find the columns with the headers
    for header in headers_to_find:
        column = find_column_by_value(sheet, header)
        if column is None and header != 'Hostname':
            raise CustomError(f'Error finding {header}')
        if column:
            header_dict[header] = column
            if return_info:
                info_to_return.append(f'{header} found in column {column}.')
            if print_info:
                print(f'{header} found in column {column}.')

    # Find the enable flag
    ssh_enabled_column = find_column_by_value(sheet, 'Multi_SSH_Enable')
    if ssh_enabled_column is None:
        raise CustomError('Error finding the enable flag')
    if return_info:
        info_to_return.append(f'Enable flag found in column {ssh_enabled_column}')
    if print_info:
        print(f'Enable flag found in column {ssh_enabled_column}')
    # Find the rows with the enable flag
    enable_row_numbers = find_rows_by_value(sheet, ssh_enabled_column, True, True)
    if not enable_row_numbers:
        raise CustomError('No enable rows found')

    # Find the positions of the commands
    numbered_header_list = ['cmd', 'end']

    command_number = 1
    numbered_header_dict = {}

    for header in numbered_header_list:
        while True:
            if command_number <= 0:
                break

            column = find_column_by_value(sheet, header + str(command_number))

            if column:
                numbered_header_dict[header + str(command_number)] = column
                if header == 'end':
                    command_number -= 1
                    continue
            else:
                command_number -= 1
                if header == 'end':
                    continue
                break
            command_number += 1
    if return_info:
        info_to_return.append(numbered_header_dict)
    if print_info:
        print(numbered_header_dict)
    if numbered_header_dict == {}:
        raise CustomError(f'Error finding {numbered_header_list[0]}')

    all_hosts = []
    for row_number in enable_row_numbers:
        single_host_dict = {'Hostname': sheet[header_dict['Hostname'] + str(row_number)].value,
                            'Address': sheet[header_dict['Address'] + str(row_number)].value,
                            'Username': sheet[header_dict['Username'] + str(row_number)].value}
        for numbered_header in numbered_header_dict:
            cell_value = sheet[numbered_header_dict[numbered_header] + str(row_number)].value
            if cell_value:
                single_host_dict[numbered_header] = cell_value
        all_hosts.append(single_host_dict)

    if return_info:
        return all_hosts, info_to_return
    return all_hosts


if __name__ == "__main__":
    result = process_excel_file(None, None, False, True)
    print(result)
    print(f'共{str(len(result))}条')
