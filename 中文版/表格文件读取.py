#!/usr/bin/python3
# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class 自定义错误(Exception):
    pass


# 在首行查找值在的列
def 根据值查找列(工作表, 值):
    for 行 in 工作表.iter_rows(min_row=1, max_row=1):
        for 单元格 in 行:
            if 单元格.value == 值:
                return 单元格.column_letter

    return None


# 在某列查找值在哪些行
def 根据列和值查找行(工作表, 列, 值, 接受空值=False):
    列索引 = column_index_from_string(列) - 1  # 转换为列索引
    行列表 = []
    for 行索引, 行 in enumerate(工作表.iter_rows(min_row=2, values_only=True), start=2):
        单元格值 = 行[列索引]
        if 单元格值 == 值:
            行列表.append(行索引)
        elif 单元格值 is None:
            if 接受空值:
                行列表.append(行索引)

    return 行列表


def 处理Excel文件(文件路径=None, 表名=None, 返回信息=False, 打印信息=False):
    if 文件路径 is None:
        文件路径 = 'SSH.xlsx'
    if 表名 is None:
        表名 = 'Sheet1'
    要返回的信息 = []

    try:
        文件 = load_workbook(文件路径)
        工作表 = 文件[表名]
    except Exception as e:
        raise 自定义错误(f'读取文件错误')

    需要查找的标题 = ['Hostname', 'Address', 'Username']
    标题字典 = {}

    # 查找标题位置
    for 标题 in 需要查找的标题:
        列 = 根据值查找列(工作表, 标题)
        if 列 is None and 标题 != 'Hostname':
            raise 自定义错误(f'查找{标题}出现错误')
        if 列:
            标题字典[标题] = 列
            if 返回信息:
                要返回的信息.append(f'{标题} 在 {列} 列.')
            if 打印信息:
                print(f'{标题} 在 {列} 列.')

    # 查找启用标记
    SSH启用标记列 = 根据值查找列(工作表, 'Multi_SSH_Enable')
    if SSH启用标记列 is None:
        raise 自定义错误(f'查找启用标记出现错误')
    if 返回信息:
        要返回的信息.append(f'启用标记在 {SSH启用标记列} 列')
    if 打印信息:
        print(f'启用标记在 {SSH启用标记列} 列')
    # 查找启用行
    启用行号列表 = 根据列和值查找行(工作表, SSH启用标记列, True, True)
    if not 启用行号列表:
        raise 自定义错误(f'没有启用行')

    # 查找命令位置
    需要查找的带序号标题列表 = ['cmd', 'end']

    命令编号 = 1
    带序号标题字典 = {}

    for 标题 in 需要查找的带序号标题列表:
        while True:
            if 命令编号 <= 0:
                break

            列 = 根据值查找列(工作表, 标题 + str(命令编号))

            if 列:
                带序号标题字典[标题 + str(命令编号)] = 列
                if 标题 == 'end':
                    命令编号 -= 1
                    continue
            else:
                命令编号 -= 1
                if 标题 == 'end':
                    continue
                break
            命令编号 += 1
    if 返回信息:
        要返回的信息.append(带序号标题字典)
    if 打印信息:
        print(带序号标题字典)
    if 带序号标题字典 == {}:
        raise 自定义错误(f'查找 {需要查找的带序号标题列表[0]} 出现错误')

    所有主机 = []
    for 行号 in 启用行号列表:
        单主机字典 = {'Hostname': 工作表[标题字典['Hostname'] + str(行号)].value,
                      'Address': 工作表[标题字典['Address'] + str(行号)].value,
                      'Username': 工作表[标题字典['Username'] + str(行号)].value}
        for 序号标题 in 带序号标题字典:
            单元格值 = 工作表[带序号标题字典[序号标题] + str(行号)].value
            if 单元格值:
                单主机字典[序号标题] = 单元格值
        所有主机.append(单主机字典)

    if 返回信息:
        return 所有主机, 要返回的信息
    return 所有主机


if __name__ == "__main__":
    结果 = 处理Excel文件(None, None, False, True)
    print(结果)
